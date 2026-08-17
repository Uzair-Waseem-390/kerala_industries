from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

DEFAULT_FILE = Path(__file__).resolve().parent.parent.parent / "customer_data" / "customer_data.xlsx"
REQUIRED_HEADERS = ["customer_code", "customer_name", "customer_address"]


class Command(BaseCommand):
    help = (
        "Imports customers from an xlsx file with columns customer_code, customer_name, "
        "customer_address (address is optional — left blank if absent). Reusable: point "
        "--file at any spreadsheet with the same headers. Safe to re-run — customers "
        "already present (matched by code) are skipped. Rows missing code or name are "
        "not imported; they're listed at the end so they can be fixed and re-run."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file", default=str(DEFAULT_FILE),
            help=f"Path to the xlsx file to import (default: {DEFAULT_FILE}).",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Report what would be created/skipped without writing anything.",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl is required for this command: pip install openpyxl")

        from billing.models import Customer

        file_path = Path(options["file"])
        dry_run = options["dry_run"]

        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        migration_user_email = getattr(settings, "DATA_GMAIL", None)
        if not migration_user_email:
            raise CommandError("DATA_GMAIL is not set in settings.py")

        User = get_user_model()
        try:
            migration_user = User.objects.get(email=migration_user_email)
        except User.DoesNotExist:
            raise CommandError(
                f"Migration attribution user '{migration_user_email}' does not exist "
                f"in the target database. Create this user first, then re-run."
            )

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(h).strip() if h is not None else "" for h in (header_row or [])]
        try:
            col = {name: headers.index(name) for name in REQUIRED_HEADERS}
        except ValueError:
            raise CommandError(
                f"Missing required column(s) in {file_path}. "
                f"Expected headers {REQUIRED_HEADERS}, found {headers}."
            )

        created = skipped = 0
        missing_rows = []

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            code = self._clean(row[col["customer_code"]]) if col["customer_code"] < len(row) else None
            name = self._clean(row[col["customer_name"]]) if col["customer_name"] < len(row) else None
            address = self._clean(row[col["customer_address"]]) if col["customer_address"] < len(row) else None

            if not code or not name:
                missing_rows.append({
                    "row": row_number, "code": code, "name": name, "address": address,
                })
                continue

            if Customer.objects.filter(code__iexact=code, is_deleted=False).exists():
                skipped += 1
                continue

            if dry_run:
                created += 1
                continue

            with transaction.atomic():
                Customer.objects.create(
                    name=name,
                    code=code.upper(),
                    address=address or "",
                    created_by=migration_user,
                    updated_by=migration_user,
                )
            created += 1

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"Import summary{' (dry run)' if dry_run else ''}: "
            f"created={created} skipped={skipped} missing={len(missing_rows)}"
        ))

        if missing_rows:
            self.stdout.write("")
            self.stdout.write(self.style.WARNING(
                f"{len(missing_rows)} row(s) skipped — missing customer_code and/or customer_name:"
            ))
            for m in missing_rows:
                self.stdout.write(
                    f"  row {m['row']}: code={m['code']!r} name={m['name']!r} address={m['address']!r}"
                )

    @staticmethod
    def _clean(value):
        if value is None:
            return None
        s = str(value).strip()
        return s if s else None
